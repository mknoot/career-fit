"""
Evaluates Nashville-area government job postings against a candidate profile.

Target fit: IT program management, digital transformation, stakeholder coordination,
            contract/procurement management, process improvement, policy work

Hard disqualifiers: PE license, CPA, ecology/field science, construction site mgmt,
HR employee relations, accounting/appropriations, emergency management (ICS/EOC),
healthcare facility ops, video production, legislative bill analysis specialist,
grant writing/federal grant lifecycle management

Edit STRONG_YES_TITLE / HARD_NO_TITLE / GREEN_FLAGS_DESC below to tune for your own profile.

Input:  nashville_jobs_full.json
Output: nashville_job_recommendations.xlsx
"""

import json
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Scoring Rules ──────────────────────────────────────────────────────────────

STRONG_YES_TITLE = [
    "program manager", "program director", "program administrator",
    "project manager", "project director",
    "it manager", "it director", "it program",
    "technology manager", "technology director", "technology program",
    "digital transformation", "information technology",
    "operations manager", "operations director",
    "contract manager", "contract administrator", "procurement manager",
    "policy analyst", "policy advisor", "policy manager",
    "portfolio manager", "portfolio director",
    "strategy", "strategic",
    "product manager", "product director",
    "data program", "data manager",
    "vendor manager", "vendor management",
    "solutions architect", "enterprise architect",
    "change management", "organizational development",
    "business analyst", "business process",
    "compliance manager", "risk manager",
    "associate director", "assistant director",
    "director of",
    "modernization", "transformation",
    "chief of staff",
    "deputy director",
    "management analyst",
    "program analyst",
    "it specialist",
    "technology specialist",
    "digital services",
    "innovation manager", "innovation director",
    "program officer",
    "implementation manager",
    "performance manager",
    "service delivery",
    "enterprise program",
    "systems program",
    # Policy/relations roles where stakeholder + govt experience transfers
    "legislative affairs", "legislative relations", "legislative liaison",
    "government relations", "public affairs",
    "intergovernmental",
]

HARD_NO_TITLE = [
    "engineer", "engineering",
    "accountant", "accounting", "auditor", "audit",
    "nurse", "physician", "doctor", "clinical", "medical", "dental", "pharmacy",
    "custodian", "maintenance", "electrician", "plumber", "hvac", "carpenter",
    "chef", "cook", "baker", "barista", "food service", "dining",
    "librarian", "archivist",
    "police", "security officer", "corrections", "detention",
    "groundskeeper", "landscaping",
    "research scientist", "postdoctoral", "postdoc",
    "professor", "lecturer", "faculty", "instructor",
    "coach", "athletic", "recreation",
    "social worker", "counselor", "therapist",
    "veterinarian", "animal",
    "painter", "driver", "transport",
    "warehouse", "inventory",
    "firefighter", "paramedic", "emt",
    # Aviation domain
    "aviation safety inspector",
    "aviation safety specialist",
    "aircrew",
    "airworthiness",
    "flight standards",
    "air traffic",
    # Operational roles where field/domain experience is required, not PM skills
    "dispatcher",          # Public Safety Dispatcher, transit dispatcher
    "transit stop",        # Transit Stop Supervisor
    "fuel lane",           # Facility Supervisor - Fuel Lane Services
    "airside",             # Operations Specialist - Airside (airport ramp ops)
    "cybersecurity administrator",  # Requires security certs, not PM background
    "corporate communications",     # PR/comms specialty, not transferable from SaaS PM
    # Medical / clinical
    "psychologist", "psychology",
    "radiolog",            # radiologist, radiology
    "audiolog",            # audiologist, audiology
    "histopatholog",       # histopathology technician
    "peer specialist",     # mental health peer support
    # Legal
    "attorney",
    # Trades / mechanical
    "mechanic",
    "boiler",
    # Law enforcement investigation
    "criminal investigator",
    # Operational support
    "parts supervisor",
    "housekeeping",
]

HARD_NO_DESC = [
    r"professional engineer|P\.E\. license|PE license",
    r"licensed professional engineer",
    r"CPA|certified public accountant",
    r"ecology|ecological|karst|endangered species",
    r"blueprint|construction site|job site|general contractor",
    r"HVAC|plumbing|electrical systems|building systems",
    r"FMLA|ADA accommodation|workplace investigation|employee relations",
    r"ICS 100|ICS 200|ICS 300|ICS 400|emergency operations center",
    r"psychiatric|inpatient psychiatric|mental health facility operations",
    r"Adobe Creative|video production|video editing|motion graphics",
    r"federal grant.*lifecycle|grant writing|CFDA|grant management",
    r"bill analysis|legislative bill|state agency rulemaking specialist",
    r"major gift|principal gift|planned giving|gift officer",
    r"fundraising|donor relations|benefactor|alumni relations",
    r"annual fund|capital campaign|endowment|development officer",
    r"prospect research|donor prospect",
    r"certified emergency manager|CEM\b",
    r"FAA certificate|pilot certificate|flight experience|airman certificate",
    r"aviation experience|aircraft.*experience|flight hours",
    # Cybersecurity specialist roles requiring technical certifications
    r"CISSP|CISM|CEH|CompTIA Security\+|security certification required",
    # Field operations requiring domain-specific licensure or hands-on experience
    r"CDL required|commercial driver|transit.*operator|bus.*operator",
    r"fuel.*handling|fueling.*procedure|airfield.*operation",
]

GREEN_FLAGS_DESC = [
    r"program management|project management",
    r"stakeholder|cross.functional",
    r"IT program|technology program|software delivery",
    r"SaaS|cloud platform|digital transformation",
    r"government.*technology|technology.*government",
    r"vendor management|contract management|procurement",
    r"policy development|policy implementation",
    r"process improvement|operational efficiency",
    r"portfolio management",
    r"change management",
    r"strategic plan|strategic initiative",
    r"federal|FedRAMP|FISMA|compliance",
    r"agile|scrum|product management",
    r"modernization|technology modernization",
    r"AI|artificial intelligence|machine learning",
    r"program delivery|program oversight|program coordination",
    r"systems integration|enterprise system|enterprise architecture",
    r"citizen services|public.facing|constituent services",
    r"data analytics|data.driven|business intelligence",
    r"cross.agency|interagency|multi.agency",
    r"performance management|performance metrics|performance measure",
    r"budget management|budget oversight|fiscal management",
    r"requirements management|business requirements|functional requirements",
    r"implementation|deployment|rollout",
    r"technology strategy|IT strategy|digital strategy",
]


# ── Scoring ────────────────────────────────────────────────────────────────────

def title_score(title: str) -> tuple[int, str]:
    t = title.lower()
    for kw in HARD_NO_TITLE:
        if kw in t:
            return -2, f"hard-no title: '{kw}'"
    for kw in STRONG_YES_TITLE:
        if kw in t:
            return 2, f"yes title: '{kw}'"
    return 0, "neutral title"


def desc_score(desc: str) -> tuple[int, str]:
    for pattern in HARD_NO_DESC:
        if re.search(pattern, desc, re.IGNORECASE):
            return -3, f"disqualifier: '{pattern}'"
    hits = [p for p in GREEN_FLAGS_DESC if re.search(p, desc, re.IGNORECASE)]
    if len(hits) >= 5:
        return 3, f"{len(hits)} green flags"
    elif len(hits) >= 3:
        return 2, f"{len(hits)} green flags"
    elif len(hits) >= 1:
        return 1, "1 green flag"
    return 0, "no strong signals"


def parse_min_salary(salary_str: str) -> float | None:
    """
    Extract the minimum salary value from a display string.
    Handles annual ($61,000) and hourly ($30.00 Hourly) formats.
    Returns an annualized float, or None if unparseable.
    """
    if not salary_str:
        return None
    s = salary_str.lower()

    # Detect hourly rates — convert to annual (2080 hrs/yr)
    hourly_match = re.search(r"\$([\d,]+(?:\.\d+)?)\s*(?:per\s*hour|hourly|/hr|/hour)", s)
    if hourly_match:
        try:
            return float(hourly_match.group(1).replace(",", "")) * 2080
        except ValueError:
            return None

    # Annual — grab the first dollar amount
    annual_match = re.search(r"\$([\d,]+(?:\.\d+)?)", s)
    if annual_match:
        try:
            return float(annual_match.group(1).replace(",", ""))
        except ValueError:
            return None

    return None


SALARY_FLOOR = 61_000  # $61k/yr or $30/hr equivalent


def evaluate(job: dict) -> dict:
    title  = job.get("title", "")
    desc   = job.get("description", "")
    salary = job.get("salary", "")

    ts, tr = title_score(title)
    ds, dr = desc_score(desc)
    total  = ts + ds

    # Salary floor — only apply when salary is explicitly known
    min_sal = parse_min_salary(salary)
    below_floor = min_sal is not None and min_sal < SALARY_FLOOR

    if below_floor:
        verdict = "NO"
    elif total >= 3:
        verdict = "YES"
    elif total == 2:
        verdict = "MAYBE-High"
    elif total == 1:
        verdict = "MAYBE-Low"
    elif total <= -2:
        verdict = "NO"
    else:
        verdict = "UNLIKELY"

    return {
        "title":            title,
        "location":         job.get("location", ""),
        "work_arrangement": get_work_arrangement(desc),
        "posted":           job.get("posted", ""),
        "salary":           salary,
        "source":           job.get("source", ""),
        "score":            total,
        "verdict":          verdict,
        "url":              job.get("url", ""),
        "reason":           f"{tr} | {dr}",
    }


def get_work_arrangement(desc: str) -> str:
    d = desc.lower()
    has_remote = bool(re.search(r"work remotely|remote work|fully remote|100% remote|telework", d))
    has_hybrid = bool(re.search(r"\bhybrid\b|flexible work arrangement|fwa", d))
    has_onsite = bool(re.search(r"on.?site.{0,30}(standard|required|only)|must.{0,30}on.?site|in.person.{0,30}required", d))

    if has_remote and has_hybrid:
        return "Hybrid"
    if has_remote and not has_onsite:
        return "Remote / Hybrid"
    if has_hybrid:
        return "Hybrid"
    if has_onsite:
        return "Onsite"
    if has_remote:
        return "Remote"
    return "Onsite"


def is_old_posting(posted: str) -> bool:
    if not posted:
        return False
    if "30+" in posted:
        return True
    m = re.search(r"(\d+)\s*[Dd]ays?\s*[Aa]go", posted)
    if m and int(m.group(1)) >= 29:
        return True
    # ISO date check (USAJOBS returns YYYY-MM-DD)
    m2 = re.match(r"(\d{4})-(\d{2})-(\d{2})", posted)
    if m2:
        from datetime import date
        try:
            post_date = date(int(m2.group(1)), int(m2.group(2)), int(m2.group(3)))
            return (date.today() - post_date).days >= 29
        except ValueError:
            pass
    return False


# ── Excel Output ───────────────────────────────────────────────────────────────

SOURCE_COLORS = {
    "Federal":          "BDD7EE",   # light blue
    "Metro Nashville":  "E2EFDA",   # light green
    "TN State":         "FFF2CC",   # light yellow
    "City of Brentwood":"D9EAD3",   # slightly darker green
    "Williamson County":"FCE5CD",   # light orange
    "MNPS":             "EAD1DC",   # light pink
    "BNA":              "D0E4F7",   # sky blue
    "WeGo":             "D9D2E9",   # light purple
    "THDA":             "F4CCCC",   # light red
    "City of Franklin": "FFE599",   # light gold
}
DEFAULT_COLOR = "F2F2F2"

HEADERS    = ["Job Title", "Source", "Salary", "Location", "Work Arrangement", "Posting Age", "Link"]
COL_WIDTHS = [50,          16,       22,       22,         18,                  20,             12]


def write_yes_sheet(wb, sheet_name: str, group: list[dict]):
    ws = wb.create_sheet(title=sheet_name)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 20

    for row_idx, r in enumerate(group, start=2):
        color = SOURCE_COLORS.get(r["source"], DEFAULT_COLOR)
        fill  = PatternFill(start_color=color, end_color=color, fill_type="solid")
        age   = r["posted"] if r["posted"] else "Unknown"
        vals  = [r["title"], r["source"], r["salary"], r["location"],
                 r["work_arrangement"], age, "Apply"]
        for col, val in enumerate(vals, start=1):
            cell           = ws.cell(row=row_idx, column=col, value=val)
            cell.fill      = fill
            cell.border    = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col == 7:
                cell.hyperlink = r["url"]
                cell.font      = Font(color="0563C1", underline="single")

    ws.freeze_panes = "A2"


def write_excel(results: list[dict], output_file: str = "nashville_job_recommendations.xlsx"):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    yes         = [r for r in results if r["verdict"] == "YES"]
    yes_recent  = [r for r in yes if not is_old_posting(r["posted"])]
    yes_old     = [r for r in yes if is_old_posting(r["posted"])]
    maybe_high  = [r for r in results if r["verdict"] == "MAYBE-High"]
    maybe_low   = [r for r in results if r["verdict"] == "MAYBE-Low"]

    write_yes_sheet(wb, "YES - Recent",       yes_recent)
    write_yes_sheet(wb, "YES - 29+ Days Old", yes_old)
    write_yes_sheet(wb, "MAYBE - High Match", maybe_high)
    write_yes_sheet(wb, "MAYBE - Low Match",  maybe_low)

    # Summary sheet
    ws = wb.create_sheet(title="Summary")
    ws["A1"] = "Nashville Metro Government Jobs"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"] = "Category"
    ws["B3"] = "Count"
    ws["A3"].font = ws["B3"].font = Font(bold=True)

    rows = [
        ("YES - Recent postings",  len(yes_recent)),
        ("YES - 29+ days old",     len(yes_old)),
        ("MAYBE - High Match",     len(maybe_high)),
        ("MAYBE - Low Match",      len(maybe_low)),
        ("UNLIKELY",               sum(1 for r in results if r["verdict"] == "UNLIKELY")),
        ("NO (filtered out)",      sum(1 for r in results if r["verdict"] == "NO")),
        ("", ""),
        ("Total evaluated",        len(results)),
        ("", ""),
        ("--- By Source ---",      ""),
    ]
    from collections import Counter
    source_counts = Counter(r["source"] for r in results if r["verdict"] == "YES")
    for src, n in sorted(source_counts.items()):
        rows.append((f"  YES from {src}", n))

    for i, (label, count) in enumerate(rows, start=4):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = count

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 10

    # Color legend
    legend_row = len(rows) + 6
    ws.cell(row=legend_row, column=1, value="Color Legend").font = Font(bold=True)
    for i, (src, color) in enumerate(SOURCE_COLORS.items(), start=legend_row + 1):
        cell = ws.cell(row=i, column=1, value=src)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    wb.save(output_file)
    print(f"Saved {output_file}")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    with open("nashville_jobs_full.json", encoding="utf-8") as f:
        jobs = json.load(f)

    results = [evaluate(j) for j in jobs]
    results.sort(key=lambda x: x["score"], reverse=True)

    yes        = [r for r in results if r["verdict"] == "YES"]
    maybe_high = [r for r in results if r["verdict"] == "MAYBE-High"]
    maybe_low  = [r for r in results if r["verdict"] == "MAYBE-Low"]
    unlikely   = [r for r in results if r["verdict"] == "UNLIKELY"]
    no         = [r for r in results if r["verdict"] == "NO"]

    print(f"Results: {len(yes)} YES | {len(maybe_high)} MAYBE-High | {len(maybe_low)} MAYBE-Low | {len(unlikely)} UNLIKELY | {len(no)} NO")

    write_excel(results)

    print("\n--- YES LIST ---\n")
    for r in yes:
        title = r["title"].encode("ascii", "replace").decode("ascii")
        sal   = f"  {r['salary']}" if r["salary"] else ""
        print(f"[{r['score']:+d}] [{r['source']}] {title}{sal}")


if __name__ == "__main__":
    main()
