"""
Evaluates UT Austin job postings against a candidate profile.

Target fit: IT program management, digital transformation, stakeholder coordination,
            contract/procurement management, process improvement, policy work

Hard disqualifiers: engineering licenses (PE), CPA, ecology/field science, construction site mgmt,
HR employee relations, accounting/appropriations, emergency management (ICS/EOC),
healthcare facility ops, video production, legislative bill analysis specialist,
requires 5+ yrs at Texas state agency, salary below $70k

Edit STRONG_YES_TITLE / HARD_NO_TITLE / GREEN_FLAGS_DESC below to tune for your own profile.
"""

import json
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- Scoring Rules ---

STRONG_YES_TITLE = [
    "program manager", "program director", "program administrator",
    "project manager", "project director",
    "IT manager", "IT director", "IT program",
    "technology manager", "technology director", "technology program",
    "digital transformation", "information technology",
    "operations manager", "operations director",
    "contract manager", "contract administrator", "procurement manager",
    "policy analyst", "policy advisor", "policy manager",
    "portfolio manager", "portfolio director",
    "strategy", "strategic",
    "product manager", "product director",
    "data program", "data manager",
    "vendor manager", "vendor management",
    "solutions architect", "enterprise architect",
    "change management", "organizational development",
    "business analyst", "business process",
    "compliance manager", "risk manager",
    "associate director", "assistant director",
    "director of",
]

HARD_NO_TITLE = [
    "engineer", "engineering",
    "accountant", "accounting", "auditor", "audit",
    "nurse", "physician", "doctor", "clinical", "medical", "dental", "pharmacy",
    "custodian", "maintenance", "electrician", "plumber", "hvac", "carpenter",
    "chef", "cook", "baker", "barista", "food service", "dining",
    "librarian", "archivist",
    "police", "security officer", "corrections",
    "groundskeeper", "landscaping",
    "research scientist", "postdoctoral", "postdoc",
    "professor", "lecturer", "faculty", "instructor",
    "coach", "athletic", "recreation",
    "social worker", "counselor", "therapist",
    "veterinarian", "animal",
    "painter", "driver", "transport",
    "warehouse", "inventory",
]

HARD_NO_DESC = [
    r"professional engineer|P\.E\. license|PE license",
    r"licensed professional engineer",
    r"CPA|certified public accountant",
    r"ecology|ecological|karst|endangered species",
    r"blueprint|construction site|job site|general contractor",
    r"HVAC|plumbing|electrical systems|building systems",
    r"FMLA|ADA accommodation|workplace investigation|employee relations",
    r"ICS 100|ICS 200|ICS 300|ICS 400|emergency operations center",
    r"psychiatric|inpatient psychiatric|mental health facility operations",
    r"Adobe Creative|video production|video editing|motion graphics",
    r"5 years.*Texas state agency|five years.*Texas state agency",
    r"federal grant.*lifecycle|grant writing|CFDA|grant management",
    r"bill analysis|legislative bill|state agency rulemaking",
    r"major gift|principal gift|planned giving|gift officer",
    r"fundraising|donor relations|benefactor|alumni relations",
    r"annual fund|capital campaign|endowment|development officer",
    r"prospect research|donor prospect",
]

GREEN_FLAGS_DESC = [
    r"program management|project management",
    r"stakeholder|cross.functional",
    r"IT program|technology program|software delivery",
    r"SaaS|cloud platform|digital transformation",
    r"government.*technology|technology.*government",
    r"vendor management|contract management|procurement",
    r"policy development|policy implementation",
    r"process improvement|operational efficiency",
    r"portfolio management",
    r"change management",
    r"strategic plan|strategic initiative",
    r"federal|FedRAMP|FISMA|compliance",
    r"agile|scrum|product management",
]


LOCATION_MAP = {
    "UT MAIN CAMPUS": "Austin, TX",
    "AUSTIN, TX": "Austin, TX",
    "AUSTIN": "Austin, TX",
    "PICKLE RESEARCH CAMPUS": "Austin, TX",
    "HOUSTON, TX": "Houston, TX",
    "SAN ANTONIO, TX": "San Antonio, TX",
    "DALLAS, TX": "Dallas, TX",
    "EL PASO, TX": "El Paso, TX",
    "ODESSA, TX": "Odessa, TX",
}

def normalize_location(location, desc):
    """Return city, state string."""
    loc = (location or "").strip().upper()
    if loc in LOCATION_MAP:
        return LOCATION_MAP[loc]
    # Try to extract from description "Location: CITY, ST"
    m = re.search(r"Location:\s*([A-Za-z ,]+,\s*[A-Z]{2})", desc)
    if m:
        return m.group(1).strip().title()
    # If it contains TX somewhere
    m2 = re.search(r"([A-Z][a-zA-Z ]+),\s*(TX|Texas)", desc)
    if m2:
        return f"{m2.group(1).strip().title()}, TX"
    if loc:
        return loc.title()
    return "Austin, TX"  # default — UT is in Austin


def get_work_arrangement(desc):
    """Detect onsite/remote/hybrid from description text."""
    d = desc.lower()

    has_remote = bool(re.search(r"work remotely|remote work|fully remote|100% remote", d))
    has_hybrid = bool(re.search(r"\bhybrid\b|flexible work arrangement|fwa|on-campus and remote", d))
    has_onsite = bool(re.search(r"on-?site.{0,30}(standard|required|expectation|only)|required.{0,30}on-?site|must.{0,30}on-?site|in.person.{0,30}required|work in.person", d))

    if has_remote and has_hybrid:
        return "Hybrid"
    if has_remote and not has_onsite:
        return "Remote / Hybrid"
    if has_hybrid:
        return "Hybrid"
    if has_onsite:
        return "Onsite"
    if has_remote:
        return "Remote"
    return "Onsite"  # UT default if not specified


def get_min_salary(desc):
    amounts = re.findall(r"\$([0-9]{1,3}(?:,[0-9]{3})+)", desc)
    nums = [int(a.replace(",", "")) for a in amounts]
    salaries = [n for n in nums if 30000 <= n <= 300000]
    return min(salaries) if salaries else None


def get_max_salary(desc):
    amounts = re.findall(r"\$([0-9]{1,3}(?:,[0-9]{3})+)", desc)
    nums = [int(a.replace(",", "")) for a in amounts]
    salaries = [n for n in nums if 30000 <= n <= 300000]
    return max(salaries) if salaries else None


def title_score(title):
    t = title.lower()
    for kw in HARD_NO_TITLE:
        if kw in t:
            return -2, f"hard-no: '{kw}'"
    for kw in STRONG_YES_TITLE:
        if kw in t:
            return 2, f"yes: '{kw}'"
    return 0, "neutral"


def desc_score(desc):
    for pattern in HARD_NO_DESC:
        if re.search(pattern, desc, re.IGNORECASE):
            return -3, f"disqualifier: '{pattern}'"
    hits = [p for p in GREEN_FLAGS_DESC if re.search(p, desc, re.IGNORECASE)]
    if len(hits) >= 4:
        return 3, f"{len(hits)} green flags"
    elif len(hits) >= 2:
        return 2, f"{len(hits)} green flags"
    elif len(hits) == 1:
        return 1, f"1 green flag"
    return 0, "no strong signals"


def evaluate(job):
    title = job.get("title", "")
    desc = job.get("description", "")
    min_sal = get_min_salary(desc)
    max_sal = get_max_salary(desc)

    ts, tr = title_score(title)
    ds, dr = desc_score(desc)
    total = ts + ds

    # Exclude sub-$70k roles (only if salary is known)
    if min_sal is not None and min_sal < 70000:
        verdict = "NO"
    elif total >= 3:
        verdict = "YES"
    elif total >= 1:
        verdict = "MAYBE"
    elif total <= -2:
        verdict = "NO"
    else:
        verdict = "UNLIKELY"

    url = f"https://utaustin.wd1.myworkdayjobs.com/en-US/UTstaff{job.get('path', '')}"

    sal_display = ""
    if min_sal and max_sal and min_sal != max_sal:
        sal_display = f"${min_sal:,} - ${max_sal:,}"
    elif min_sal:
        sal_display = f"${min_sal:,}"

    return {
        "title": title,
        "location": normalize_location(job.get("location", ""), desc),
        "work_arrangement": get_work_arrangement(desc),
        "posted": job.get("posted", ""),
        "salary": sal_display,
        "min_salary": min_sal,
        "score": total,
        "verdict": verdict,
        "url": url,
        "reason": f"{tr} | {dr}",
    }


def is_old_posting(posted):
    """Returns True if posted 29+ days ago."""
    if not posted:
        return False
    if "30+" in posted:
        return True
    m = re.search(r"(\d+) Days Ago", posted, re.IGNORECASE)
    if m and int(m.group(1)) >= 29:
        return True
    return False


def write_yes_sheet(wb, sheet_name, group, color_hex):
    headers = ["Job Title", "Salary Range", "Location", "Work Arrangement", "Posting Age", "Link"]
    col_widths = [50, 22, 18, 18, 20, 12]

    ws = wb.create_sheet(title=sheet_name)
    fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 20

    for row_idx, r in enumerate(group, start=2):
        age_label = r["posted"] if r["posted"] else "Unknown"
        vals = [r["title"], r["salary"], r["location"], r["work_arrangement"], age_label, "Apply"]
        for col, val in enumerate(vals, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col == 6:
                cell.hyperlink = r["url"]
                cell.font = Font(color="0563C1", underline="single")

    ws.freeze_panes = "A2"


def write_excel(results):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    yes = [r for r in results if r["verdict"] == "YES"]
    yes_recent = [r for r in yes if not is_old_posting(r["posted"])]
    yes_old = [r for r in yes if is_old_posting(r["posted"])]

    write_yes_sheet(wb, "YES - Recent", yes_recent, "92D050")   # green
    write_yes_sheet(wb, "YES - 29+ Days Old", yes_old, "FFD7D7")  # light red

    # Summary sheet
    ws_sum = wb.create_sheet(title="Summary")
    ws_sum["A1"] = "UT Austin Job Evaluation"
    ws_sum["A1"].font = Font(bold=True, size=14)
    ws_sum["A3"] = "Category"
    ws_sum["B3"] = "Count"
    ws_sum["A3"].font = Font(bold=True)
    ws_sum["B3"].font = Font(bold=True)
    rows = [
        ("YES - Recent postings", len(yes_recent)),
        ("YES - 29+ days old", len(yes_old)),
        ("MAYBE", sum(1 for r in results if r["verdict"] == "MAYBE")),
        ("UNLIKELY", sum(1 for r in results if r["verdict"] == "UNLIKELY")),
        ("NO (filtered out)", sum(1 for r in results if r["verdict"] == "NO")),
        ("Total evaluated", len(results)),
    ]
    for i, (label, count) in enumerate(rows, start=4):
        ws_sum[f"A{i}"] = label
        ws_sum[f"B{i}"] = count
    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 10

    wb.save("job_recommendations.xlsx")
    print("Saved job_recommendations.xlsx")


def main():
    with open("jobs_full.json", encoding="utf-8") as f:
        jobs = json.load(f)

    results = [evaluate(j) for j in jobs]
    results.sort(key=lambda x: x["score"], reverse=True)

    yes = [r for r in results if r["verdict"] == "YES"]
    maybe = [r for r in results if r["verdict"] == "MAYBE"]
    no = [r for r in results if r["verdict"] == "NO"]
    unlikely = [r for r in results if r["verdict"] == "UNLIKELY"]

    print(f"Results: {len(yes)} YES | {len(maybe)} MAYBE | {len(unlikely)} UNLIKELY | {len(no)} NO")

    write_excel(results)

    print("\n--- YES LIST ---\n")
    for r in yes:
        title = r["title"].encode("ascii", "replace").decode("ascii")
        sal = f"  {r['salary']}" if r["salary"] else "  (no salary listed)"
        print(f"[{r['score']:+d}] {title}{sal}")


if __name__ == "__main__":
    main()
